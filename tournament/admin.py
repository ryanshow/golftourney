import io
from datetime import date
from fpdf import FPDF
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from django.contrib import admin
from django.http import HttpResponse, Http404
from django.urls import path, reverse
from django.utils.html import format_html

from .models import SponsorshipPackage, Sponsor, Registration, RegistrationPlayer, TournamentInfo, AddOn, RegistrationAddOn, RaffleDonation


@admin.register(TournamentInfo)
class TournamentInfoAdmin(admin.ModelAdmin):
    fieldsets = (
        ('Tournament Details', {
            'fields': ('tournament_name', 'date', 'location', 'description', 'entry_deadline'),
        }),
        ('Contact Information', {
            'fields': ('contact_name', 'contact_email', 'contact_phone'),
        }),
    )

    def has_add_permission(self, request):
        # Only allow adding if no instance exists
        if TournamentInfo.objects.exists():
            return False
        return True

    def has_delete_permission(self, request, obj=None):
        return False


@admin.register(SponsorshipPackage)
class SponsorshipPackageAdmin(admin.ModelAdmin):
    list_display = ['name', 'price', 'max_players', 'sponsor_tier', 'logo_upload', 'is_active', 'sort_order']
    list_editable = ['sponsor_tier', 'logo_upload', 'is_active', 'sort_order']
    list_filter = ['is_active', 'logo_upload', 'sponsor_tier']
    ordering = ['sort_order', 'price']


@admin.register(Sponsor)
class SponsorAdmin(admin.ModelAdmin):
    list_display = ['name', 'tier', 'is_active', 'sort_order', 'logo_preview']
    list_editable = ['is_active', 'tier', 'sort_order']
    list_filter = ['tier', 'is_active']
    ordering = ['tier', 'sort_order', 'name']

    def logo_preview(self, obj):
        if obj.logo:
            return format_html(
                '<img src="{}" style="height:40px; width:auto; object-fit:contain;" />',
                obj.logo.url
            )
        return "No logo"
    logo_preview.short_description = "Logo Preview"


@admin.register(AddOn)
class AddOnAdmin(admin.ModelAdmin):
    list_display = ['name', 'price', 'is_active', 'sort_order']
    list_editable = ['price', 'is_active', 'sort_order']
    ordering = ['sort_order', 'name']


class RegistrationAddOnInline(admin.TabularInline):
    model = RegistrationAddOn
    extra = 0
    readonly_fields = ['addon', 'addon_price']
    can_delete = False

    def addon_price(self, obj):
        return f"${obj.addon.price}"
    addon_price.short_description = "Price"


class RegistrationPlayerInline(admin.TabularInline):
    model = RegistrationPlayer
    extra = 0
    fields = ['slot', 'name']
    readonly_fields = ['slot']
    can_delete = False
    verbose_name = "Player"
    verbose_name_plural = "Players in Group"


@admin.register(Registration)
class RegistrationAdmin(admin.ModelAdmin):
    change_list_template = 'admin/tournament/registration/change_list.html'

    list_display = [
        'full_name', 'email', 'sponsorship_package', 'payment_method',
        'payment_status', 'order_total', 'has_logo', 'logo_approved', 'created_at',
        'invoice_link',
    ]
    list_editable = ['logo_approved']
    list_filter = ['payment_method', 'payment_status', 'sponsorship_package', 'logo_approved', 'created_at']
    search_fields = ['first_name', 'last_name', 'email', 'company_org', 'square_payment_id']
    readonly_fields = ['token', 'created_at', 'square_payment_id', 'logo_preview']
    date_hierarchy = 'created_at'
    actions = ['approve_logos', 'revoke_logo_approval']

    inlines = [RegistrationPlayerInline, RegistrationAddOnInline]
    fieldsets = (
        ('Registrant Information', {
            'fields': ('first_name', 'last_name', 'email', 'phone', 'company_org'),
        }),
        ('Sponsorship', {
            'fields': ('sponsorship_package',),
        }),
        ('Payment', {
            'fields': ('payment_method', 'payment_status', 'square_payment_id'),
        }),
        ('Logo', {
            'fields': ('company_logo', 'logo_preview', 'logo_approved'),
        }),
        ('Additional', {
            'fields': ('notes', 'created_at', 'token'),
        }),
    )

    @admin.action(description="Approve selected logos for display on the home page")
    def approve_logos(self, request, queryset):
        updated = queryset.filter(company_logo__isnull=False).exclude(company_logo='').update(logo_approved=True)
        self.message_user(request, f"{updated} logo(s) approved.")

    @admin.action(description="Revoke logo approval (hide from home page)")
    def revoke_logo_approval(self, request, queryset):
        updated = queryset.update(logo_approved=False)
        self.message_user(request, f"{updated} logo approval(s) revoked.")

    def get_urls(self):
        urls = super().get_urls()
        custom = [
            path('export-excel/', self.admin_site.admin_view(self.export_excel), name='registration_export_excel'),
            path('<int:pk>/invoice-pdf/', self.admin_site.admin_view(self.invoice_pdf), name='registration_invoice_pdf'),
        ]
        return custom + urls

    def export_excel(self, request):
        registrations = (
            Registration.objects
            .select_related('sponsorship_package')
            .prefetch_related('players', 'addons__addon')
            .order_by('sponsorship_package__sort_order', 'last_name', 'first_name')
        )

        # Determine max player count across all registrations
        max_players = max(
            (reg.sponsorship_package.max_players for reg in registrations),
            default=0
        )

        wb = Workbook()
        ws = wb.active
        ws.title = "Registrations"

        # Styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="1D4582")
        center = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # Build header row
        headers = [
            'First Name', 'Last Name', 'Email', 'Phone',
            'Company / Organization', 'Sponsorship Package',
            'Add-Ons', 'Add-On Total', 'Package Price', 'Order Total',
            'Payment Method', 'Paid?',
        ]
        for i in range(1, max_players + 1):
            headers.append(f'Player {i}')

        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center

        # Data rows
        for row_idx, reg in enumerate(registrations, start=2):
            addon_qs = reg.addons.select_related('addon').all()
            addon_names = ', '.join(ra.addon.name for ra in addon_qs)
            addon_total = sum(ra.addon.price for ra in addon_qs)
            order_total = reg.sponsorship_package.price + addon_total

            players_by_slot = {p.slot: p.name for p in reg.players.all()}

            row = [
                reg.first_name,
                reg.last_name,
                reg.email,
                reg.phone,
                reg.company_org,
                reg.sponsorship_package.name,
                addon_names,
                float(addon_total),
                float(reg.sponsorship_package.price),
                float(order_total),
                reg.get_payment_method_display(),
                'Yes' if reg.payment_status == 'paid' else 'No',
            ]
            for slot in range(1, max_players + 1):
                row.append(players_by_slot.get(slot, ''))

            for col, value in enumerate(row, start=1):
                cell = ws.cell(row=row_idx, column=col, value=value)
                cell.alignment = Alignment(vertical="center", wrap_text=True)

            # Alternate row shading
            if row_idx % 2 == 0:
                fill = PatternFill("solid", fgColor="EEF2F9")
                for col in range(1, len(headers) + 1):
                    ws.cell(row=row_idx, column=col).fill = fill

        # Format currency columns
        money_cols = [headers.index('Add-On Total') + 1, headers.index('Package Price') + 1, headers.index('Order Total') + 1]
        for col in money_cols:
            for row_idx in range(2, ws.max_row + 1):
                ws.cell(row=row_idx, column=col).number_format = '"$"#,##0.00'

        # Auto-fit column widths
        for col in ws.columns:
            max_len = max((len(str(cell.value or '')) for cell in col), default=10)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 40)

        ws.freeze_panes = 'A2'

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        response = HttpResponse(
            buf,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename="registrations.xlsx"'
        return response

    def invoice_pdf(self, request, pk):
        try:
            reg = (
                Registration.objects
                .select_related('sponsorship_package')
                .prefetch_related('addons__addon', 'players')
                .get(pk=pk)
            )
        except Registration.DoesNotExist:
            raise Http404

        tournament = TournamentInfo.get_instance()
        addon_qs = reg.addons.select_related('addon').all()
        addon_total = sum(ra.addon.price for ra in addon_qs)
        order_total = reg.sponsorship_package.price + addon_total

        # --- Colours & layout constants ---
        NAVY   = (29, 69, 130)   # #1D4582
        LIGHT  = (238, 242, 249)
        WHITE  = (255, 255, 255)
        BLACK  = (0, 0, 0)
        GRAY   = (100, 100, 100)
        GREEN  = (40, 120, 60)
        RED    = (180, 40, 40)
        MARGIN = 15
        PAGE_W = 210  # A4 width in mm

        class InvoicePDF(FPDF):
            def header(self):
                pass
            def footer(self):
                self.set_y(-12)
                self.set_font("Helvetica", "I", 8)
                self.set_text_color(*GRAY)
                self.cell(0, 5, f"Page {self.page_no()}", align="C")

        pdf = InvoicePDF(orientation="P", unit="mm", format="A4")
        pdf.set_auto_page_break(auto=True, margin=20)
        pdf.add_page()
        pdf.set_margins(MARGIN, MARGIN, MARGIN)

        # ── Header bar ──────────────────────────────────────────────────
        pdf.set_fill_color(*NAVY)
        pdf.rect(0, 0, PAGE_W, 28, "F")
        pdf.set_y(6)
        pdf.set_font("Helvetica", "B", 16)
        pdf.set_text_color(*WHITE)
        pdf.cell(0, 8, tournament.tournament_name or "Golf Tournament", align="C", new_x="LMARGIN", new_y="NEXT")
        pdf.set_font("Helvetica", "", 9)
        pdf.set_text_color(200, 220, 255)
        sub_parts = []
        if tournament.date:
            sub_parts.append(tournament.date.strftime("%B %-d, %Y"))
        if tournament.location:
            sub_parts.append(tournament.location)
        if sub_parts:
            pdf.cell(0, 6, "  |  ".join(sub_parts), align="C", new_x="LMARGIN", new_y="NEXT")

        # Ensure we're clearly below the navy banner before the INVOICE title
        pdf.set_y(max(pdf.get_y(), 28) + 10)

        # ── "INVOICE" title + meta ───────────────────────────────────────
        pdf.set_text_color(*NAVY)
        pdf.set_font("Helvetica", "B", 22)
        pdf.cell(0, 10, "INVOICE", new_x="LMARGIN", new_y="NEXT")
        pdf.set_font("Helvetica", "", 9)
        pdf.set_text_color(*GRAY)
        pdf.cell(0, 5, f"Ref: {str(reg.token).upper()}   |   Date: {reg.created_at.strftime('%B %-d, %Y')}", new_x="LMARGIN", new_y="NEXT")
        pdf.ln(4)

        # ── Bill To / From columns ───────────────────────────────────────
        col_w = (PAGE_W - 2 * MARGIN) / 2

        # Headers
        pdf.set_font("Helvetica", "B", 9)
        pdf.set_text_color(*NAVY)
        pdf.cell(col_w, 6, "BILL TO", new_x="RIGHT", new_y="TOP")
        pdf.set_x(MARGIN + col_w)
        pdf.cell(col_w, 6, "FROM", new_x="LMARGIN", new_y="NEXT")

        pdf.set_font("Helvetica", "", 9)
        pdf.set_text_color(*BLACK)

        bill_lines = [reg.full_name]
        if reg.company_org:
            bill_lines.append(reg.company_org)
        bill_lines.append(reg.email)
        if reg.phone:
            bill_lines.append(reg.phone)

        from_lines = ["Clay Elementary Parent Club", "12449 S. Smith Ave.", "Kingsburg, CA 93631"]
        if tournament.contact_email:
            from_lines.append(tournament.contact_email)
        if tournament.contact_phone:
            from_lines.append(tournament.contact_phone)

        max_lines = max(len(bill_lines), len(from_lines))
        for i in range(max_lines):
            bill_text = bill_lines[i] if i < len(bill_lines) else ""
            from_text = from_lines[i] if i < len(from_lines) else ""
            pdf.cell(col_w, 5, bill_text, new_x="RIGHT", new_y="TOP")
            pdf.set_x(MARGIN + col_w)
            pdf.cell(col_w, 5, from_text, new_x="LMARGIN", new_y="NEXT")

        pdf.ln(6)

        # ── Itemised table ───────────────────────────────────────────────
        usable_w = PAGE_W - 2 * MARGIN
        desc_w   = usable_w * 0.72
        amt_w    = usable_w * 0.28

        # Table header
        pdf.set_fill_color(*NAVY)
        pdf.set_text_color(*WHITE)
        pdf.set_font("Helvetica", "B", 9)
        pdf.cell(desc_w, 7, "  DESCRIPTION", fill=True, new_x="RIGHT", new_y="TOP")
        pdf.cell(amt_w,  7, "AMOUNT", fill=True, align="R", new_x="LMARGIN", new_y="NEXT")

        # Sponsorship package row
        pdf.set_fill_color(*LIGHT)
        pdf.set_text_color(*BLACK)
        pdf.set_font("Helvetica", "", 9)
        pdf.cell(desc_w, 7, f"  {reg.sponsorship_package.name}", fill=True, new_x="RIGHT", new_y="TOP")
        pdf.cell(amt_w,  7, f"${reg.sponsorship_package.price:,.2f}", align="R", fill=True, new_x="LMARGIN", new_y="NEXT")

        # Package benefits (indented, small)
        pdf.set_font("Helvetica", "I", 8)
        pdf.set_text_color(*GRAY)
        for benefit in reg.sponsorship_package.benefits_list():
            pdf.cell(desc_w, 5, f"      - {benefit}", new_x="RIGHT", new_y="TOP")
            pdf.cell(amt_w,  5, "", new_x="LMARGIN", new_y="NEXT")

        # Add-on rows
        pdf.set_font("Helvetica", "", 9)
        pdf.set_text_color(*BLACK)
        for i, ra in enumerate(addon_qs):
            fill_color = WHITE if i % 2 == 0 else LIGHT
            pdf.set_fill_color(*fill_color)
            pdf.cell(desc_w, 7, f"  {ra.addon.name} (Add-On)", fill=True, new_x="RIGHT", new_y="TOP")
            pdf.cell(amt_w,  7, f"${ra.addon.price:,.2f}", align="R", fill=True, new_x="LMARGIN", new_y="NEXT")

        # Divider
        pdf.set_draw_color(*NAVY)
        pdf.set_line_width(0.5)
        pdf.line(MARGIN, pdf.get_y(), PAGE_W - MARGIN, pdf.get_y())
        pdf.ln(1)

        # Total row
        pdf.set_font("Helvetica", "B", 10)
        pdf.set_fill_color(*NAVY)
        pdf.set_text_color(*WHITE)
        pdf.cell(desc_w, 8, "  TOTAL", fill=True, new_x="RIGHT", new_y="TOP")
        pdf.cell(amt_w,  8, f"${order_total:,.2f}", align="R", fill=True, new_x="LMARGIN", new_y="NEXT")
        pdf.ln(6)

        # ── Payment status ───────────────────────────────────────────────
        if reg.payment_status == 'paid':
            status_color = GREEN
            status_note  = f"Payment received via {reg.get_payment_method_display()}."
            if reg.square_payment_id:
                status_note += f"  Ref: {reg.square_payment_id}"
        elif reg.payment_status == 'pending':
            status_color = (180, 120, 0)
            if reg.payment_method == 'check':
                status_note = (
                    "Payment due by check.\n"
                    "Make check payable to:\n"
                    "  Clay Elementary Parent Club\n"
                    "Mail to:\n"
                    "  Clay Elementary Parent Club\n"
                    "  12449 S. Smith Ave.\n"
                    "  Kingsburg, CA 93631\n"
                    f"Memo: Golf Tournament - {reg.full_name}"
                )
            else:
                status_note = "Payment pending."
        else:
            status_color = RED
            status_note  = "Payment failed or was not completed."

        status_label = reg.get_payment_status_display().upper()
        badge_w = 40
        note_w  = usable_w - badge_w
        pdf.set_fill_color(*status_color)
        pdf.set_text_color(*WHITE)
        pdf.set_font("Helvetica", "B", 10)
        badge_y = pdf.get_y()
        pdf.cell(badge_w, 8, f"  {status_label}", fill=True, new_x="RIGHT", new_y="TOP")
        pdf.set_font("Helvetica", "", 8)
        pdf.set_text_color(*GRAY)
        pdf.set_x(MARGIN + badge_w)
        pdf.multi_cell(note_w, 4.5, f"  {status_note}", new_x="LMARGIN", new_y="NEXT")
        pdf.ln(5)

        # ── Players ──────────────────────────────────────────────────────
        players = list(reg.players.all())
        if players:
            pdf.set_font("Helvetica", "B", 9)
            pdf.set_text_color(*NAVY)
            pdf.cell(0, 6, "PLAYERS IN GROUP", new_x="LMARGIN", new_y="NEXT")
            pdf.set_line_width(0.3)
            pdf.set_draw_color(*NAVY)
            pdf.line(MARGIN, pdf.get_y(), PAGE_W - MARGIN, pdf.get_y())
            pdf.ln(2)
            pdf.set_font("Helvetica", "", 9)
            pdf.set_text_color(*BLACK)
            for p in players:
                name = p.name if p.name else "TBD"
                pdf.cell(0, 5, f"  {p.slot}.  {name}", new_x="LMARGIN", new_y="NEXT")
            pdf.ln(4)

        # ── Tax / EIN notice ─────────────────────────────────────────────
        pdf.set_fill_color(*LIGHT)
        pdf.set_draw_color(*NAVY)
        pdf.set_line_width(0.3)
        notice_y = pdf.get_y()
        notice_h = 16
        pdf.rect(MARGIN, notice_y, usable_w, notice_h, "FD")
        pdf.set_xy(MARGIN + 3, notice_y + 2)
        pdf.set_font("Helvetica", "B", 8)
        pdf.set_text_color(*NAVY)
        pdf.cell(0, 4, "TAX DEDUCTIBILITY NOTICE", new_x="LMARGIN", new_y="NEXT")
        pdf.set_x(MARGIN + 3)
        pdf.set_font("Helvetica", "", 8)
        pdf.set_text_color(*GRAY)
        pdf.multi_cell(
            usable_w - 6, 4,
            "Clay Elementary Parent Club is a 501(c)(3) nonprofit organization (EIN: 46-2668474). "
            "Your sponsorship payment may be tax deductible as provided by law. "
            "Please retain this invoice for your records.",
            new_x="LMARGIN", new_y="NEXT",
        )

        buf = io.BytesIO(bytes(pdf.output()))
        response = HttpResponse(buf, content_type="application/pdf")
        safe_name = f"invoice-{str(reg.token)[:8]}-{reg.last_name.lower()}.pdf"
        response["Content-Disposition"] = f'inline; filename="{safe_name}"'
        return response

    def invoice_link(self, obj):
        url = reverse("admin:registration_invoice_pdf", args=[obj.pk])
        return format_html(
            '<a href="{}" target="_blank" style="'
            'background:#1d4582;color:#fff;padding:3px 10px;border-radius:3px;'
            'text-decoration:none;font-size:11px;white-space:nowrap;">'
            '&#128196; PDF Invoice</a>',
            url,
        )
    invoice_link.short_description = "Invoice"

    def full_name(self, obj):
        return obj.full_name
    full_name.short_description = "Name"
    full_name.admin_order_field = 'last_name'

    def order_total(self, obj):
        addon_total = sum(ra.addon.price for ra in obj.addons.select_related('addon').all())
        total = obj.sponsorship_package.price + addon_total
        return f"${total:,.2f}"
    order_total.short_description = "Order Total"

    def has_logo(self, obj):
        if not obj.company_logo:
            return None  # renders as dash
        if obj.logo_approved:
            return True   # green checkmark
        return False      # red X = uploaded but pending approval
    has_logo.short_description = "Logo"
    has_logo.boolean = True

    def logo_preview(self, obj):
        if obj.company_logo:
            return format_html(
                '<img src="{}" style="max-height:80px; max-width:200px; object-fit:contain;" />',
                obj.company_logo.url
            )
        return "No logo uploaded"
    logo_preview.short_description = "Logo Preview"


@admin.register(RaffleDonation)
class RaffleDonationAdmin(admin.ModelAdmin):
    list_display = ['full_name', 'email', 'phone', 'company_name', 'estimated_value', 'delivery_method', 'created_at']
    search_fields = ['first_name', 'last_name', 'email', 'company_name']
    list_filter = ['delivery_method']
    readonly_fields = ['created_at']
    ordering = ['-created_at']

    fieldsets = (
        ('Donor Information', {
            'fields': ('first_name', 'last_name', 'email', 'phone', 'company_name'),
        }),
        ('Donation', {
            'fields': ('donation_description', 'estimated_value', 'delivery_method', 'created_at'),
        }),
    )

    def full_name(self, obj):
        return obj.full_name
    full_name.short_description = "Name"
    full_name.admin_order_field = 'last_name'
